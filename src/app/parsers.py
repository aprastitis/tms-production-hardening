import csv
import io
from dataclasses import dataclass
from typing import List, Optional

from openpyxl import load_workbook


@dataclass
class RawRow:
    row_number: int
    payload: dict
    error: Optional[str] = None


def _parse_csv(data: bytes) -> List[RawRow]:
    text = data.decode("utf-8", errors="replace")
    reader = csv.DictReader(io.StringIO(text))
    rows = []
    for i, row in enumerate(reader, start=1):
        rows.append(RawRow(row_number=i, payload=dict(row)))
    return rows


def _parse_xlsx(data: bytes) -> List[RawRow]:
    wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)
    try:
        header = [str(c) if c is not None else "" for c in next(rows_iter)]
    except StopIteration:
        return []
    rows = []
    for i, vals in enumerate(rows_iter, start=1):
        payload = {header[j]: vals[j] for j in range(min(len(header), len(vals))) if vals[j] is not None}
        rows.append(RawRow(row_number=i, payload=payload))
    return rows


def _parse_mt940(data: bytes) -> List[RawRow]:
    text = data.decode("utf-8", errors="replace")
    rows: List[RawRow] = []
    current = {"reference": "", "amount": "", "currency": "", "value_date": "", "counterparty": ""}
    i = 0
    lines = text.splitlines()
    while i < len(lines):
        line = lines[i].strip()
        if line.startswith(":61:"):
            parts = line[4:].split("N", 1)
            if len(parts) == 2:
                # YYMMDD + amount
                amt = parts[1].lstrip("0123456789,").replace(",", ".")
                current["amount"] = amt
        elif line.startswith(":86:"):
            current["counterparty"] = line[4:].strip()[:128]
            rows.append(RawRow(row_number=len(rows) + 1, payload=dict(current)))
            current = {"reference": "", "amount": "", "currency": "", "value_date": "", "counterparty": ""}
        i += 1
    return rows


def _parse_bai2(data: bytes) -> List[RawRow]:
    text = data.decode("utf-8", errors="replace")
    rows: List[RawRow] = []
    for i, line in enumerate(text.splitlines(), start=1):
        if not line or line[0] not in ("16",):
            continue
        parts = line.split(",")
        if len(parts) >= 4:
            rows.append(RawRow(row_number=i, payload={"type_code": parts[0], "amount": parts[2], "reference": parts[3] if len(parts) > 3 else ""}))
    return rows


def dispatch(data: bytes, content_type: str, source) -> List[RawRow]:
    ct = (content_type or "").lower()
    if "csv" in ct or source.parser_kind == "CSV":
        return _parse_csv(data)
    if "excel" in ct or "spreadsheet" in ct or source.parser_kind == "XLSX":
        return _parse_xlsx(data)
    if source.parser_kind == "MT940":
        return _parse_mt940(data)
    if source.parser_kind == "BAI2":
        return _parse_bai2(data)
    # Default CSV
    return _parse_csv(data)